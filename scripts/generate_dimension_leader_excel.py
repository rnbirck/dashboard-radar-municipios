from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as WorksheetImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from generate_dimension_leader_radars import (
    COLORS,
    EXPORT_ROOT,
    YEAR,
    LeaderRadar,
    build_leader_radars,
)


OUTPUT_PATH = (
    EXPORT_ROOT
    / "municipios-lideres-por-dimensao-rf1-rf9-2025.xlsx"
)

EXCEL_COLORS = {
    "ink": COLORS["ink"].removeprefix("#"),
    "muted": COLORS["muted"].removeprefix("#"),
    "border": COLORS["border"].removeprefix("#"),
    "header": "E9F2F6",
    "teal": COLORS["teal"].removeprefix("#"),
    "teal_fill": "DDEFEA",
    "orange": COLORS["orange"].removeprefix("#"),
    "orange_fill": "FFF0D7",
    "row_alt": "F8FAFC",
    "white": "FFFFFF",
    "rank_fill": "E9EEF3",
    "rank_text": "344A5F",
}

THIN = Side(style="thin", color=EXCEL_COLORS["border"])
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(cell, *, left: bool = False) -> None:
    cell.fill = PatternFill("solid", fgColor=EXCEL_COLORS["header"])
    cell.font = Font(
        name="Aptos Display",
        size=11,
        bold=True,
        color=EXCEL_COLORS["ink"],
    )
    cell.alignment = Alignment(
        horizontal="left" if left else "center",
        vertical="center",
        wrap_text=True,
    )
    cell.border = CELL_BORDER


def style_body(
    cell,
    alternate: bool,
    *,
    left: bool = False,
    bold: bool = False,
    muted: bool = False,
) -> None:
    cell.fill = PatternFill(
        "solid",
        fgColor=(
            EXCEL_COLORS["row_alt"]
            if alternate
            else EXCEL_COLORS["white"]
        ),
    )
    cell.font = Font(
        name="Aptos",
        size=10.5,
        bold=bold,
        color=(
            EXCEL_COLORS["muted"]
            if muted
            else EXCEL_COLORS["ink"]
        ),
    )
    cell.alignment = Alignment(
        horizontal="left" if left else "center",
        vertical="center",
        wrap_text=True,
    )
    cell.border = CELL_BORDER


def style_rank(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=EXCEL_COLORS["rank_fill"])
    cell.font = Font(
        name="Aptos",
        size=10.5,
        bold=True,
        color=EXCEL_COLORS["rank_text"],
    )
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '0"º"'
    cell.border = CELL_BORDER


def style_score(cell, *, comparison: bool = False) -> None:
    cell.fill = PatternFill(
        "solid",
        fgColor=(
            EXCEL_COLORS["orange_fill"]
            if comparison
            else EXCEL_COLORS["teal_fill"]
        ),
    )
    cell.font = Font(
        name="Aptos",
        size=10.5,
        bold=True,
        color=(
            EXCEL_COLORS["orange"]
            if comparison
            else EXCEL_COLORS["teal"]
        ),
    )
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0.00"
    cell.border = CELL_BORDER


def add_summary_sheet(
    workbook: Workbook,
    records: list[LeaderRadar],
) -> None:
    worksheet = workbook.create_sheet("Resumo")
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.zoomScale = 90

    headers = [
        "RF",
        "DIMENSÃO",
        "MUNICÍPIO LÍDER",
        "COREDE",
        "POSIÇÃO 2025",
        "NOTA DA DIMENSÃO",
        "INDICADORES",
        "ARQUIVO PNG",
    ]
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column, value=header)
        style_header(cell, left=column in (2, 3, 4, 8))
    worksheet.row_dimensions[1].height = 34

    for row_index, record in enumerate(records, start=2):
        alternate = row_index % 2 == 1
        values = [
            f"RF{record.region_number}",
            record.dimension_name,
            record.municipality_name,
            record.corede_name,
            record.dimension_rank,
            record.dimension_score,
            len(record.indicators),
            record.output_path.name,
        ]
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_index, column=column, value=value)
            if column == 5:
                style_rank(cell)
            elif column == 6:
                style_score(cell)
            else:
                style_body(
                    cell,
                    alternate,
                    left=column in (2, 3, 4, 8),
                    bold=column in (1, 2, 3),
                    muted=column in (4, 8),
                )
        worksheet.cell(row=row_index, column=8).hyperlink = (
            record.output_path.resolve().as_uri()
        )
        worksheet.cell(row=row_index, column=8).style = "Hyperlink"
        worksheet.row_dimensions[row_index].height = 29

    worksheet.auto_filter.ref = f"A1:H{len(records) + 1}"
    widths = [10, 22, 30, 30, 16, 20, 14, 54]
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width

    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.oddHeader.center.text = (
        "Municípios líderes por dimensão — Regiões Funcionais — 2025"
    )
    worksheet.oddHeader.center.font = "Aptos,Bold"
    worksheet.oddFooter.right.text = "Página &P de &N"


def add_region_sheet(
    workbook: Workbook,
    region_number: int,
    records: list[LeaderRadar],
) -> None:
    region_records = [
        record for record in records if record.region_number == region_number
    ]
    worksheet = workbook.create_sheet(f"RF{region_number}")
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 75
    worksheet.freeze_panes = "A12"

    worksheet.merge_cells("A1:J1")
    title_cell = worksheet["A1"]
    title_cell.value = (
        f"RF{region_number} — municípios líderes por dimensão — {YEAR}"
    )
    title_cell.fill = PatternFill(
        "solid",
        fgColor=EXCEL_COLORS["header"],
    )
    title_cell.font = Font(
        name="Aptos Display",
        size=14,
        bold=True,
        color=EXCEL_COLORS["ink"],
    )
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    title_cell.border = CELL_BORDER
    worksheet.row_dimensions[1].height = 32

    summary_headers = [
        "DIMENSÃO",
        "MUNICÍPIO LÍDER",
        "COREDE",
        "POSIÇÃO",
        "NOTA DA DIMENSÃO",
        "INDICADORES",
        "ARQUIVO PNG",
    ]
    for column, header in enumerate(summary_headers, start=1):
        cell = worksheet.cell(row=3, column=column, value=header)
        style_header(cell, left=column in (1, 2, 3, 7))
    worksheet.row_dimensions[3].height = 34

    for row_index, record in enumerate(region_records, start=4):
        alternate = row_index % 2 == 1
        values = [
            record.dimension_name,
            record.municipality_name,
            record.corede_name,
            record.dimension_rank,
            record.dimension_score,
            len(record.indicators),
            record.output_path.name,
        ]
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_index, column=column, value=value)
            if column == 4:
                style_rank(cell)
            elif column == 5:
                style_score(cell)
            else:
                style_body(
                    cell,
                    alternate,
                    left=column in (1, 2, 3, 7),
                    bold=column in (1, 2),
                    muted=column in (3, 7),
                )
        worksheet.cell(row=row_index, column=7).hyperlink = (
            record.output_path.resolve().as_uri()
        )
        worksheet.cell(row=row_index, column=7).style = "Hyperlink"
        worksheet.row_dimensions[row_index].height = 31

    detail_headers = [
        "DIMENSÃO",
        "MUNICÍPIO",
        "INDICADOR",
        "NOTA MUNICÍPIO",
        "MEDIANA RF",
        "DIFERENÇA",
        "VALOR ORIGINAL",
        "MEDIANA ORIGINAL RF",
        "RANK DO INDICADOR",
        "AMOSTRA DA MEDIANA",
    ]
    detail_header_row = 12
    for column, header in enumerate(detail_headers, start=1):
        cell = worksheet.cell(
            row=detail_header_row,
            column=column,
            value=header,
        )
        style_header(cell, left=column in (1, 2, 3))
    worksheet.row_dimensions[detail_header_row].height = 40

    score_comment = Comment(
        "Notas padronizadas na escala de 0 a 10, usadas no radar.",
        "Codex",
    )
    worksheet.cell(detail_header_row, 4).comment = score_comment
    worksheet.cell(detail_header_row, 5).comment = score_comment

    detail_row = detail_header_row + 1
    for record in region_records:
        for indicator in record.indicators:
            municipality_score = indicator["score"]
            median_score = indicator["regional_median_score"]
            difference = (
                float(municipality_score) - float(median_score)
                if municipality_score is not None
                and median_score is not None
                else None
            )
            values = [
                record.dimension_name,
                record.municipality_name,
                indicator["indicator_name"],
                municipality_score,
                median_score,
                difference,
                indicator["original_value"],
                indicator["regional_median_original_value"],
                indicator["indicator_rank"],
                indicator["regional_median_sample_size"],
            ]
            alternate = detail_row % 2 == 0
            for column, value in enumerate(values, start=1):
                cell = worksheet.cell(
                    row=detail_row,
                    column=column,
                    value=value,
                )
                if column == 4:
                    style_score(cell)
                elif column == 5:
                    style_score(cell, comparison=True)
                elif column == 9:
                    style_rank(cell)
                else:
                    style_body(
                        cell,
                        alternate,
                        left=column in (1, 2, 3),
                        bold=column in (1, 2),
                        muted=column in (7, 8, 10),
                    )
                    if column in (6, 7, 8):
                        cell.number_format = "0.0000"
            worksheet.row_dimensions[detail_row].height = 34
            detail_row += 1

    worksheet.auto_filter.ref = (
        f"A{detail_header_row}:J{detail_row - 1}"
    )

    widths = [20, 29, 48, 18, 17, 15, 21, 24, 20, 22]
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width
    for column in range(11, 31):
        worksheet.column_dimensions[get_column_letter(column)].width = 9

    image_anchors = ("L2", "U2", "L27", "U27", "L52", "U52")
    for record, anchor in zip(region_records, image_anchors):
        image = WorksheetImage(record.output_path)
        image.width = 675
        image.height = 450
        worksheet.add_image(image, anchor)

    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.35
    worksheet.page_margins.bottom = 0.35
    worksheet.oddHeader.center.text = (
        f"RF{region_number} — líderes por dimensão — {YEAR}"
    )
    worksheet.oddHeader.center.font = "Aptos,Bold"
    worksheet.oddFooter.right.text = "Página &P de &N"


def validate_workbook(
    path: Path,
    records: list[LeaderRadar],
) -> None:
    workbook = load_workbook(path, data_only=False)
    expected_sheets = ["Resumo", *[f"RF{number}" for number in range(1, 10)]]
    if workbook.sheetnames != expected_sheets:
        raise ValueError(
            f"Abas inesperadas: {workbook.sheetnames}; "
            f"esperado: {expected_sheets}"
        )

    summary = workbook["Resumo"]
    if summary.max_row != len(records) + 1 or summary.max_column != 8:
        raise ValueError(
            f"Resumo com tamanho inesperado: "
            f"{summary.max_row}x{summary.max_column}"
        )
    if summary["C2"].value != records[0].municipality_name:
        raise ValueError("Resumo não inicia com o primeiro líder esperado")

    for region_number in range(1, 10):
        worksheet = workbook[f"RF{region_number}"]
        region_records = [
            record
            for record in records
            if record.region_number == region_number
        ]
        if len(region_records) != 6:
            raise ValueError(
                f"RF{region_number}: esperadas 6 dimensões, "
                f"encontradas {len(region_records)}"
            )
        if len(worksheet._images) != 6:
            raise ValueError(
                f"RF{region_number}: esperadas 6 imagens, "
                f"encontradas {len(worksheet._images)}"
            )
        expected_detail_rows = sum(
            len(record.indicators) for record in region_records
        )
        actual_detail_rows = worksheet.max_row - 12
        if actual_detail_rows != expected_detail_rows:
            raise ValueError(
                f"RF{region_number}: {actual_detail_rows} linhas de "
                f"indicadores; esperado {expected_detail_rows}"
            )

        for offset, record in enumerate(region_records, start=4):
            if worksheet.cell(offset, 2).value != record.municipality_name:
                raise ValueError(
                    f"RF{region_number} {record.dimension_name}: "
                    "município divergente no resumo"
                )
            if worksheet.cell(offset, 4).value != 1:
                raise ValueError(
                    f"RF{region_number} {record.dimension_name}: "
                    "posição diferente de 1"
                )

    workbook.close()


def main() -> None:
    records = build_leader_radars()
    missing_images = [
        record.output_path
        for record in records
        if not record.output_path.exists()
    ]
    if missing_images:
        raise FileNotFoundError(
            "Radares ausentes. Execute primeiro "
            "generate_dimension_leader_radars.py: "
            + ", ".join(str(path) for path in missing_images[:3])
        )

    EXPORT_ROOT.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = (
        "Municípios líderes por dimensão nas Regiões Funcionais — 2025"
    )
    workbook.properties.subject = (
        "Líder de cada dimensão, indicadores, notas e medianas regionais"
    )
    workbook.properties.creator = "Codex"
    workbook.properties.description = (
        "Resumo dos 54 líderes dimensionais e dados utilizados nos radares."
    )

    add_summary_sheet(workbook, records)
    for region_number in range(1, 10):
        add_region_sheet(workbook, region_number, records)

    workbook.save(OUTPUT_PATH)
    validate_workbook(OUTPUT_PATH, records)
    print(f"Excel gerado e validado: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
