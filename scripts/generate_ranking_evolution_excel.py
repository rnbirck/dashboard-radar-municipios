from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_ROOT = PROJECT_ROOT / "public" / "data" / "v2025" / "rankings"
EXPORT_ROOT = PROJECT_ROOT / "docs" / "exports"
OUTPUT_PATH = (
    EXPORT_ROOT
    / "evolucao-ranking-geral-regioes-funcionais-2021-2025.xlsx"
)
YEARS = tuple(range(2021, 2026))

COLORS = {
    "ink": "142A41",
    "muted": "52657A",
    "header": "E9F2F6",
    "border": "D8E2E9",
    "rank_fill": "E9EEF3",
    "rank_text": "344A5F",
    "row_alt": "F8FAFC",
    "white": "FFFFFF",
}

THIN = Side(style="thin", color=COLORS["border"])
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_region(year: int, region_number: int) -> dict:
    path = DATA_ROOT / str(year) / f"rf{region_number}.json"
    return json.loads(path.read_text(encoding="utf-8"))["data"]


def style_header(cell, *, left: bool = False) -> None:
    cell.fill = PatternFill("solid", fgColor=COLORS["header"])
    cell.font = Font(
        name="Aptos Display",
        size=12,
        bold=True,
        color=COLORS["ink"],
    )
    cell.alignment = Alignment(
        horizontal="left" if left else "center",
        vertical="center",
        wrap_text=True,
    )
    cell.border = CELL_BORDER


def style_identity(cell, alternate: bool, *, muted: bool = False) -> None:
    cell.fill = PatternFill(
        "solid",
        fgColor=COLORS["row_alt"] if alternate else COLORS["white"],
    )
    cell.font = Font(
        name="Aptos",
        size=12,
        bold=True,
        color=COLORS["muted"] if muted else COLORS["ink"],
    )
    cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )
    cell.border = CELL_BORDER


def style_rank(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=COLORS["rank_fill"])
    cell.font = Font(
        name="Aptos",
        size=12,
        bold=True,
        color=COLORS["rank_text"],
    )
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '0"º"'
    cell.border = CELL_BORDER


def add_region_sheet(workbook: Workbook, region_number: int) -> None:
    annual_data = {
        year: load_region(year, region_number) for year in YEARS
    }
    top_2025 = annual_data[2025]["municipalities"][:5]
    annual_by_id = {
        year: {
            municipality["municipalityId"]: municipality
            for municipality in annual_data[year]["municipalities"]
        }
        for year in YEARS
    }

    worksheet = workbook.create_sheet(title=f"RF{region_number}")
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 90
    worksheet.freeze_panes = "D2"
    worksheet.auto_filter.ref = "A1:H6"

    headers = [
        "POS. 2025",
        "MUNICÍPIO",
        "COREDE",
        *[str(year) for year in YEARS],
    ]
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column, value=header)
        style_header(cell, left=column in (2, 3))
    worksheet.row_dimensions[1].height = 34

    for row_index, municipality in enumerate(top_2025, start=2):
        alternate = row_index % 2 == 1
        worksheet.row_dimensions[row_index].height = 36

        final_rank = worksheet.cell(
            row=row_index,
            column=1,
            value=int(municipality["overallRank"]),
        )
        municipality_cell = worksheet.cell(
            row=row_index,
            column=2,
            value=municipality["municipalityName"],
        )
        corede_cell = worksheet.cell(
            row=row_index,
            column=3,
            value=municipality["coredeName"],
        )
        style_rank(final_rank)
        style_identity(municipality_cell, alternate)
        style_identity(corede_cell, alternate, muted=True)

        for year_offset, year in enumerate(YEARS, start=4):
            historical = annual_by_id[year][municipality["municipalityId"]]
            rank_cell = worksheet.cell(
                row=row_index,
                column=year_offset,
                value=int(historical["overallRank"]),
            )
            style_rank(rank_cell)

    widths = {
        "A": 14,
        "B": 28,
        "C": 30,
        "D": 12,
        "E": 12,
        "F": 12,
        "G": 12,
        "H": 12,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    chart_path = (
        EXPORT_ROOT
        / f"evolucao-ranking-geral-rf{region_number}-2021-2025.png"
    )
    if not chart_path.exists():
        raise FileNotFoundError(f"Gráfico ausente: {chart_path}")

    chart_image = Image(chart_path)
    original_width = chart_image.width
    original_height = chart_image.height
    chart_image.width = 1200
    chart_image.height = round(
        original_height * chart_image.width / original_width
    )
    worksheet.add_image(chart_image, "A9")

    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.35
    worksheet.page_margins.bottom = 0.35
    worksheet.oddHeader.center.text = (
        f"Região Funcional {region_number} — evolução do ranking geral"
    )
    worksheet.oddHeader.center.font = "Aptos,Bold"
    worksheet.oddHeader.center.size = 11
    worksheet.oddFooter.right.text = "Página &P de &N"


def validate_workbook(path: Path) -> None:
    workbook = load_workbook(path, data_only=False)
    expected_sheets = [f"RF{region_number}" for region_number in range(1, 10)]
    if workbook.sheetnames != expected_sheets:
        raise ValueError(
            f"Abas inesperadas: {workbook.sheetnames}; "
            f"esperado: {expected_sheets}"
        )

    for region_number, sheet_name in enumerate(expected_sheets, start=1):
        worksheet = workbook[sheet_name]
        if worksheet.max_row != 6 or worksheet.max_column != 8:
            raise ValueError(
                f"{sheet_name}: tamanho inesperado "
                f"{worksheet.max_row}x{worksheet.max_column}"
            )
        if worksheet["A1"].value != "POS. 2025":
            raise ValueError(f"{sheet_name}: cabeçalho inicial inválido")
        if worksheet["H1"].value != "2025":
            raise ValueError(f"{sheet_name}: último ano inválido")
        if len(worksheet._images) != 1:
            raise ValueError(
                f"{sheet_name}: esperado 1 gráfico incorporado, "
                f"encontrados {len(worksheet._images)}"
            )

        current = load_region(2025, region_number)
        expected_names = [
            municipality["municipalityName"]
            for municipality in current["municipalities"][:5]
        ]
        actual_names = [
            worksheet.cell(row=row, column=2).value for row in range(2, 7)
        ]
        if actual_names != expected_names:
            raise ValueError(
                f"{sheet_name}: municípios divergentes: {actual_names}"
            )

        for row, municipality in enumerate(
            current["municipalities"][:5],
            start=2,
        ):
            municipality_id = municipality["municipalityId"]
            for column, year in enumerate(YEARS, start=4):
                annual = load_region(year, region_number)
                historical = next(
                    item
                    for item in annual["municipalities"]
                    if item["municipalityId"] == municipality_id
                )
                expected_rank = int(historical["overallRank"])
                actual_rank = worksheet.cell(row=row, column=column).value
                if actual_rank != expected_rank:
                    raise ValueError(
                        f"{sheet_name} {municipality['municipalityName']} "
                        f"{year}: {actual_rank} != {expected_rank}"
                    )

    workbook.close()


def main() -> None:
    EXPORT_ROOT.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = (
        "Evolução do ranking geral das Regiões Funcionais — 2021 a 2025"
    )
    workbook.properties.subject = (
        "Cinco primeiros municípios de 2025 em cada RF"
    )
    workbook.properties.creator = "Codex"
    workbook.properties.description = (
        "Posições anuais de 2021 a 2025 e gráficos incorporados para RF1 a RF9."
    )

    for region_number in range(1, 10):
        add_region_sheet(workbook, region_number)

    workbook.save(OUTPUT_PATH)
    validate_workbook(OUTPUT_PATH)
    print(f"Excel gerado e validado: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
