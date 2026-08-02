from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_ROOT = PROJECT_ROOT / "public" / "data" / "v2025" / "rankings"
OUTPUT_PATH = (
    PROJECT_ROOT
    / "docs"
    / "exports"
    / "ranking-regioes-funcionais-2025-vs-2024.xlsx"
)

DIMENSIONS = [
    ("GERAL", "overallRank"),
    ("EDUCAÇÃO", "educacao"),
    ("FINANÇAS", "financas"),
    ("MEIO AMBIENTE", "meioAmbiente"),
    ("SAÚDE", "saude"),
    ("SEGURANÇA", "seguranca"),
    ("SOCIOECONÔMICO", "socioeconomico"),
]

COLORS = {
    "ink": "142A41",
    "muted": "52657A",
    "border": "D8E2E9",
    "border_strong": "BDCCD8",
    "header": "E9F2F6",
    "row_alt": "FBFCFD",
    "white": "FFFFFF",
    "rank_fill": "E9EEF3",
    "rank_border": "C7D1DB",
    "rank_text": "344A5F",
    "positive_fill": "DFF3EA",
    "positive_border": "B7DFCF",
    "positive_text": "087654",
    "negative_fill": "FDE8E8",
    "negative_border": "EFC1C4",
    "negative_text": "B4232D",
    "neutral_fill": "EEF2F6",
    "neutral_border": "D7E0E7",
    "neutral_text": "526277",
}

THIN_BORDER = Side(style="thin", color=COLORS["border"])
MEDIUM_BORDER = Side(style="medium", color=COLORS["border_strong"])


def load_region(year: int, region_number: int) -> dict:
    path = DATA_ROOT / str(year) / f"rf{region_number}.json"
    return json.loads(path.read_text(encoding="utf-8"))["data"]


def get_rank(municipality: dict, key: str) -> int:
    if key == "overallRank":
        return int(municipality["overallRank"])
    return int(municipality["dimensionRanks"][key])


def style_group_header(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=COLORS["header"])
    cell.font = Font(name="Aptos Display", size=11, bold=True, color=COLORS["ink"])
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=MEDIUM_BORDER,
        right=MEDIUM_BORDER,
        top=MEDIUM_BORDER,
        bottom=THIN_BORDER,
    )


def style_subheader(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=COLORS["header"])
    cell.font = Font(name="Aptos", size=10, bold=True, color=COLORS["muted"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=THIN_BORDER,
        right=THIN_BORDER,
        top=THIN_BORDER,
        bottom=MEDIUM_BORDER,
    )


def style_identity_cell(cell, alternate: bool) -> None:
    cell.fill = PatternFill(
        "solid", fgColor=COLORS["row_alt"] if alternate else COLORS["white"]
    )
    cell.font = Font(name="Aptos", size=11, bold=True, color=COLORS["ink"])
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = Border(
        left=THIN_BORDER,
        right=THIN_BORDER,
        top=THIN_BORDER,
        bottom=THIN_BORDER,
    )


def style_rank_cell(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=COLORS["rank_fill"])
    cell.font = Font(name="Aptos", size=11, bold=True, color=COLORS["rank_text"])
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '0"º"'
    cell.border = Border(
        left=Side(style="thin", color=COLORS["rank_border"]),
        right=Side(style="thin", color=COLORS["rank_border"]),
        top=Side(style="thin", color=COLORS["rank_border"]),
        bottom=Side(style="thin", color=COLORS["rank_border"]),
    )


def style_variation_cell(cell, value: int) -> None:
    if value > 0:
        fill = COLORS["positive_fill"]
        border = COLORS["positive_border"]
        text = COLORS["positive_text"]
    elif value < 0:
        fill = COLORS["negative_fill"]
        border = COLORS["negative_border"]
        text = COLORS["negative_text"]
    else:
        fill = COLORS["neutral_fill"]
        border = COLORS["neutral_border"]
        text = COLORS["neutral_text"]

    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name="Aptos", size=11, bold=True, color=text)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '+0;-0;0'
    cell.border = Border(
        left=Side(style="thin", color=border),
        right=Side(style="thin", color=border),
        top=Side(style="thin", color=border),
        bottom=Side(style="thin", color=border),
    )


def add_region_sheet(workbook: Workbook, region_number: int) -> None:
    current = load_region(2025, region_number)
    previous = load_region(2024, region_number)
    previous_by_id = {
        municipality["municipalityId"]: municipality
        for municipality in previous["municipalities"]
    }

    worksheet = workbook.create_sheet(title=f"RF{region_number}")
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "C3"
    worksheet.auto_filter.ref = "A2:P17"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.print_title_rows = "1:2"
    worksheet.print_area = "A1:P17"
    worksheet.sheet_properties.outlinePr.summaryBelow = False

    worksheet.oddHeader.center.text = (
        f"Região Funcional {region_number} — ranking 2025 e variação frente a 2024"
    )
    worksheet.oddHeader.center.size = 11
    worksheet.oddHeader.center.font = "Aptos,Bold"
    worksheet.oddFooter.right.text = "Página &P de &N"
    worksheet.oddFooter.right.size = 9

    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.45
    worksheet.page_margins.bottom = 0.45
    worksheet.page_margins.header = 0.2
    worksheet.page_margins.footer = 0.2

    worksheet.merge_cells("A1:B1")
    worksheet["A1"] = "IDENTIFICAÇÃO"
    style_group_header(worksheet["A1"])

    start_column = 3
    for dimension_index, (label, _) in enumerate(DIMENSIONS):
        group_start = start_column + dimension_index * 2
        group_end = group_start + 1
        worksheet.merge_cells(
            start_row=1,
            start_column=group_start,
            end_row=1,
            end_column=group_end,
        )
        header_cell = worksheet.cell(row=1, column=group_start, value=label)
        style_group_header(header_cell)

    subheaders = ["MUNICÍPIO", "COREDE"]
    for _ in DIMENSIONS:
        subheaders.extend(["POS. 2025", "VAR."])

    variation_comment = Comment(
        "Variação = posição em 2024 menos posição em 2025. "
        "Valor positivo indica melhora no ranking; valor negativo indica queda.",
        "Codex",
    )

    for column, value in enumerate(subheaders, start=1):
        cell = worksheet.cell(row=2, column=column, value=value)
        style_subheader(cell)
        if value == "VAR.":
            cell.comment = variation_comment

    worksheet.row_dimensions[1].height = 25
    worksheet.row_dimensions[2].height = 30

    top_municipalities = current["municipalities"][:15]
    for row_offset, municipality in enumerate(top_municipalities, start=3):
        previous_municipality = previous_by_id[municipality["municipalityId"]]
        alternate = row_offset % 2 == 0
        worksheet.row_dimensions[row_offset].height = 32

        municipality_cell = worksheet.cell(
            row=row_offset, column=1, value=municipality["municipalityName"]
        )
        corede_cell = worksheet.cell(
            row=row_offset, column=2, value=municipality["coredeName"]
        )
        style_identity_cell(municipality_cell, alternate)
        style_identity_cell(corede_cell, alternate)
        corede_cell.font = Font(
            name="Aptos", size=10, bold=True, color=COLORS["muted"]
        )

        for dimension_index, (_, key) in enumerate(DIMENSIONS):
            rank = get_rank(municipality, key)
            previous_rank = get_rank(previous_municipality, key)
            variation = previous_rank - rank

            rank_column = 3 + dimension_index * 2
            variation_column = rank_column + 1

            rank_cell = worksheet.cell(
                row=row_offset, column=rank_column, value=rank
            )
            variation_cell = worksheet.cell(
                row=row_offset, column=variation_column, value=variation
            )
            style_rank_cell(rank_cell)
            style_variation_cell(variation_cell, variation)

    column_widths = {
        "A": 25,
        "B": 25,
    }
    for column in range(3, 17):
        letter = get_column_letter(column)
        column_widths[letter] = 12 if column % 2 == 1 else 13

    for letter, width in column_widths.items():
        worksheet.column_dimensions[letter] = ColumnDimension(
            worksheet, index=letter, width=width
        )

    for column in (1, 2):
        worksheet.cell(row=2, column=column).alignment = Alignment(
            horizontal="left", vertical="center"
        )

    worksheet.auto_filter.ref = "A2:P17"


def validate_workbook(path: Path) -> None:
    workbook = load_workbook(path, data_only=False)
    expected_sheets = [f"RF{index}" for index in range(1, 10)]
    if workbook.sheetnames != expected_sheets:
        raise ValueError(
            f"Abas inesperadas: {workbook.sheetnames}; esperado: {expected_sheets}"
        )

    for sheet_name in expected_sheets:
        worksheet = workbook[sheet_name]
        if worksheet.max_row != 17 or worksheet.max_column != 16:
            raise ValueError(
                f"{sheet_name}: tamanho inesperado "
                f"{worksheet.max_row}x{worksheet.max_column}"
            )
        if worksheet["A2"].value != "MUNICÍPIO":
            raise ValueError(f"{sheet_name}: cabeçalho de município inválido")
        if worksheet["P2"].value != "VAR.":
            raise ValueError(f"{sheet_name}: última coluna inválida")
        municipalities = [
            worksheet.cell(row=row, column=1).value for row in range(3, 18)
        ]
        if len(set(municipalities)) != 15:
            raise ValueError(f"{sheet_name}: municípios ausentes ou duplicados")

    workbook.close()


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = (
        "Ranking das Regiões Funcionais — 2025 versus 2024"
    )
    workbook.properties.subject = (
        "Top 15 municípios por Região Funcional, ranking geral e dimensões"
    )
    workbook.properties.creator = "Codex"
    workbook.properties.description = (
        "Posições de 2025 e variações frente a 2024 para RF1 a RF9."
    )

    for region_number in range(1, 10):
        add_region_sheet(workbook, region_number)

    workbook.save(OUTPUT_PATH)
    validate_workbook(OUTPUT_PATH)
    print(f"Excel gerado e validado: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
