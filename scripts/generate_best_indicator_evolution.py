from __future__ import annotations

import argparse
import json
import math
import shutil
import textwrap
import unicodedata
import zipfile
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

from generate_dimension_leader_radars import (
    CATALOG_PATH,
    MUNICIPALITY_ROOT,
    LeaderRadar,
    build_leader_radars,
    load_json,
)


PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXPORT_ROOT = PROJECT_ROOT / "docs" / "exports" / "graficos 5"
CHART_ROOT = EXPORT_ROOT / "evolucao-indicadores"
WORKBOOK_PATH = (
    EXPORT_ROOT
    / "melhores-indicadores-lideres-dimensionais-rf1-rf9-2021-2025.xlsx"
)
ZIP_PATH = EXPORT_ROOT / "graficos-e-excel-melhores-indicadores-rf1-rf9.zip"

REFERENCE_YEARS = tuple(range(2021, 2026))
CANVAS_WIDTH = 3200
CANVAS_HEIGHT = 1900
DPI = 300

REGULAR_FONT_PATH = Path(r"C:\Windows\Fonts\segoeui.ttf")
SEMIBOLD_FONT_PATH = Path(r"C:\Windows\Fonts\seguisb.ttf")
BOLD_FONT_PATH = Path(r"C:\Windows\Fonts\segoeuib.ttf")

COLORS = {
    "ink": "#142A41",
    "muted": "#52657A",
    "muted_light": "#68798B",
    "border": "#D8E2E9",
    "plot_border": "#E1E9EE",
    "plot": "#F8FBFC",
    "grid": "#D7E2E8",
    "teal": "#08716D",
    "teal_dark": "#075B58",
    "teal_area": (8, 106, 102, 28),
    "orange": "#B86F12",
    "orange_dark": "#80500E",
    "state": "#536B85",
    "header": "#E9F2F6",
    "row_alt": "#F8FAFC",
    "white": "#FFFFFF",
    "rank_fill": "#E9EEF3",
    "rank_text": "#344A5F",
}


@dataclass(frozen=True)
class EvolutionPoint:
    data_year: int
    reference_year: int
    municipality_value: float | None
    regional_median: float | None
    state_median: float | None
    indicator_rank: int | None
    is_imputed: bool | None


@dataclass(frozen=True)
class IndicatorEvolution:
    region_number: int
    region_id: str
    dimension_id: str
    dimension_name: str
    municipality_id: str
    municipality_name: str
    corede_name: str
    dimension_rank: int
    dimension_score: float | None
    indicator_id: str
    indicator_name: str
    indicator_short_name: str
    indicator_rank_2025: int
    indicator_score_2025: float | None
    description: str
    direction: str
    unit: str | None
    value_format: str | None
    decimal_places: int
    multiplier: float
    points: tuple[EvolutionPoint, ...]
    output_path: Path


def font(size: int, *, bold: bool = False, semibold: bool = False) -> ImageFont.FreeTypeFont:
    if bold:
        path = BOLD_FONT_PATH
    elif semibold:
        path = SEMIBOLD_FONT_PATH
    else:
        path = REGULAR_FONT_PATH
    return ImageFont.truetype(str(path), size=size)


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    cleaned = "".join(
        character.lower() if character.isalnum() else "-"
        for character in ascii_value
    )
    return "-".join(part for part in cleaned.split("-") if part)


def coalesce(*values: Any) -> Any:
    return next((value for value in values if value is not None), None)


def pt_number(
    value: float,
    *,
    maximum_digits: int = 1,
    minimum_digits: int = 0,
) -> str:
    rendered = f"{value:,.{maximum_digits}f}"
    integer, dot, fractional = rendered.partition(".")
    if maximum_digits:
        fractional = fractional.rstrip("0")
        if len(fractional) < minimum_digits:
            fractional = fractional.ljust(minimum_digits, "0")
    integer = integer.replace(",", "\u0000").replace(".", ",").replace("\u0000", ".")
    return f"{integer},{fractional}" if fractional else integer


def scaled_value(value: float, record: IndicatorEvolution) -> float:
    return value * record.multiplier


def format_point_value(value: float | None, record: IndicatorEvolution) -> str:
    if value is None or not math.isfinite(value):
        return "—"
    display_value = scaled_value(value, record)
    minimum_digits = record.decimal_places if record.indicator_id == "vinculos_per_capita" else 0
    rendered = pt_number(
        display_value,
        maximum_digits=record.decimal_places,
        minimum_digits=minimum_digits,
    )
    if record.value_format == "currency":
        return f"R$ {rendered}"
    if record.value_format == "percent" or record.unit == "%":
        return f"{rendered}%"
    return rendered


def format_full_value(value: float | None, record: IndicatorEvolution) -> str:
    rendered = format_point_value(value, record)
    if value is None or rendered == "—":
        return rendered
    if record.value_format in ("currency", "percent") or record.unit == "%":
        return rendered
    return f"{rendered} {record.unit}" if record.unit else rendered


def axis_label(record: IndicatorEvolution) -> str:
    if record.indicator_id == "saeb_ensino_fundamental":
        return "Nota"
    if record.value_format == "percent" or record.unit == "%":
        return "Percentual (%)"
    if record.value_format == "currency":
        return f"Valor ({record.unit})" if record.unit else "Valor (R$)"
    return record.unit or record.indicator_short_name or "Valor do indicador"


def format_axis_tick(value: float, record: IndicatorEvolution) -> str:
    scaled = scaled_value(value, record)
    magnitude = abs(scaled)
    suffix = ""
    divisor = 1.0
    if magnitude >= 1_000_000_000:
        divisor, suffix = 1_000_000_000, " bi"
    elif magnitude >= 1_000_000:
        divisor, suffix = 1_000_000, " mi"
    elif magnitude >= 1_000:
        divisor, suffix = 1_000, " mil"
    if suffix:
        return f"{pt_number(scaled / divisor, maximum_digits=1)}{suffix}"
    return format_point_value(value, record)


def direction_text(direction: str) -> str:
    if direction == "higher_is_better":
        return "Valores mais altos indicam melhor desempenho."
    if direction == "lower_is_better":
        return "Valores mais baixos indicam melhor desempenho."
    return "Direção interpretativa neutra."


def selection_rule_text(rank: int) -> str:
    if rank == 1:
        return "1º lugar no indicador em 2025"
    return f"melhor posição do município no indicador em 2025: {rank}º"


def build_evolution_points(
    *,
    indicator: dict[str, Any],
    metadata: dict[str, Any],
    region_id: str,
    reference_years: Sequence[int],
) -> tuple[EvolutionPoint, ...]:
    rows_by_reference_year = {
        int(item["year"]): item for item in indicator["values"]
    }
    regional_catalog = (
        metadata
        .get("regionalMedianOriginalValueByRegionAndReferenceYear", {})
        .get(region_id, {})
    )
    state_catalog = metadata.get("stateMedianOriginalValueByReferenceYear", {})
    data_year_mapping = metadata.get("dataYearByReferenceYear", {})
    points_by_data_year: dict[int, EvolutionPoint] = {}

    for reference_year in sorted(reference_years):
        row = rows_by_reference_year.get(reference_year)
        data_year = int(
            data_year_mapping.get(str(reference_year), reference_year)
        )
        current = points_by_data_year.get(data_year)
        regional_value = regional_catalog.get(str(reference_year))
        own_regional_value = (
            row.get("regionalMedianOriginalValue") if row else None
        )
        state_value = state_catalog.get(str(reference_year))
        points_by_data_year[data_year] = EvolutionPoint(
            data_year=data_year,
            reference_year=reference_year,
            municipality_value=coalesce(
                row.get("originalValue") if row else None,
                current.municipality_value if current else None,
            ),
            regional_median=coalesce(
                regional_value,
                own_regional_value,
                current.regional_median if current else None,
            ),
            state_median=coalesce(
                state_value,
                current.state_median if current else None,
            ),
            indicator_rank=coalesce(
                row.get("rank") if row else None,
                current.indicator_rank if current else None,
            ),
            is_imputed=coalesce(
                row.get("isImputed") if row else None,
                current.is_imputed if current else None,
            ),
        )
    return tuple(points_by_data_year[year] for year in sorted(points_by_data_year))


def build_records(
    *,
    region_numbers: Iterable[int] = range(1, 10),
) -> list[IndicatorEvolution]:
    leaders = build_leader_radars(region_numbers=region_numbers)
    catalog = load_json(CATALOG_PATH)
    metadata_by_id = {
        item["id"]: item for item in catalog["indicators"]
    }
    records: list[IndicatorEvolution] = []

    for leader in leaders:
        available_ranks = [
            int(item["indicator_rank"])
            for item in leader.indicators
            if item.get("indicator_rank") is not None
        ]
        if not available_ranks:
            raise ValueError(
                f"{leader.region_id} {leader.dimension_name} "
                f"{leader.municipality_name}: nenhum indicador ranqueado em 2025"
            )
        best_indicator_rank = min(available_ranks)
        selected_ids = {
            item["indicator_id"]
            for item in leader.indicators
            if item.get("indicator_rank") == best_indicator_rank
        }
        dimension_data = load_json(
            MUNICIPALITY_ROOT
            / leader.municipality_id
            / f"{leader.dimension_id}.json"
        )
        indicators_by_id = {
            item["indicatorId"]: item
            for item in dimension_data["indicators"]
        }

        for indicator_id in sorted(
            selected_ids,
            key=lambda value: (
                metadata_by_id[value].get("order", 999),
                metadata_by_id[value].get("name", value),
            ),
        ):
            metadata = metadata_by_id[indicator_id]
            indicator = indicators_by_id[indicator_id]
            current = next(
                item
                for item in indicator["values"]
                if int(item["year"]) == 2025
            )
            indicator_name = metadata.get("name") or indicator_id
            output_path = (
                CHART_ROOT
                / f"rf{leader.region_number}"
                / (
                    f"evolucao-rf{leader.region_number}-"
                    f"{slugify(leader.dimension_name)}-"
                    f"{slugify(leader.municipality_name)}-"
                    f"{slugify(indicator_name)}.png"
                )
            )
            description = (
                "Recursos de acesso à informação nas escolas."
                if indicator_id == "qt_acesso_infor"
                else metadata.get("description")
                or "Descrição metodológica não disponível."
            )
            records.append(
                IndicatorEvolution(
                    region_number=leader.region_number,
                    region_id=leader.region_id,
                    dimension_id=leader.dimension_id,
                    dimension_name=leader.dimension_name,
                    municipality_id=leader.municipality_id,
                    municipality_name=leader.municipality_name,
                    corede_name=leader.corede_name,
                    dimension_rank=leader.dimension_rank,
                    dimension_score=leader.dimension_score,
                    indicator_id=indicator_id,
                    indicator_name=indicator_name,
                    indicator_short_name=metadata.get("shortName")
                    or indicator_name,
                    indicator_rank_2025=int(current["rank"]),
                    indicator_score_2025=current.get("score"),
                    description=description,
                    direction=metadata.get("direction") or "neutral",
                    unit=metadata.get("unit"),
                    value_format=metadata.get("format"),
                    decimal_places=int(metadata.get("decimalPlaces") or 0),
                    multiplier=float(metadata.get("multiplier") or 1),
                    points=build_evolution_points(
                        indicator=indicator,
                        metadata=metadata,
                        region_id=leader.region_id,
                        reference_years=tuple(
                            int(year)
                            for year in dimension_data["availableYears"]
                            if int(year) in REFERENCE_YEARS
                        ),
                    ),
                    output_path=output_path,
                )
            )

    return sorted(
        records,
        key=lambda item: (
            item.region_number,
            item.dimension_id,
            item.municipality_name,
            item.indicator_name,
        ),
    )


def text_width(draw: ImageDraw.ImageDraw, value: str, text_font: ImageFont.FreeTypeFont) -> float:
    return draw.textlength(value, font=text_font)


def wrap_by_pixels(
    draw: ImageDraw.ImageDraw,
    value: str,
    text_font: ImageFont.FreeTypeFont,
    max_width: int,
) -> list[str]:
    words = value.split()
    if not words:
        return [""]
    lines: list[str] = []
    current = words[0]
    for word in words[1:]:
        candidate = f"{current} {word}"
        if text_width(draw, candidate, text_font) <= max_width:
            current = candidate
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return lines


def draw_dashed_line(
    draw: ImageDraw.ImageDraw,
    points: Sequence[tuple[float, float]],
    *,
    fill: str,
    width: int,
    dash: tuple[float, float],
) -> None:
    if len(points) < 2:
        return
    dash_length, gap_length = dash
    for start, end in zip(points, points[1:]):
        x1, y1 = start
        x2, y2 = end
        dx, dy = x2 - x1, y2 - y1
        length = math.hypot(dx, dy)
        if length <= 0:
            continue
        cursor = 0.0
        while cursor < length:
            segment_end = min(cursor + dash_length, length)
            ratio_start = cursor / length
            ratio_end = segment_end / length
            draw.line(
                (
                    x1 + dx * ratio_start,
                    y1 + dy * ratio_start,
                    x1 + dx * ratio_end,
                    y1 + dy * ratio_end,
                ),
                fill=fill,
                width=width,
            )
            cursor += dash_length + gap_length


def series_segments(
    points: Sequence[EvolutionPoint],
    getter,
    x_values: Sequence[float],
    y_value,
) -> list[list[tuple[float, float]]]:
    segments: list[list[tuple[float, float]]] = []
    active: list[tuple[float, float]] = []
    for index, point in enumerate(points):
        value = getter(point)
        if value is None:
            if active:
                segments.append(active)
                active = []
            continue
        active.append((x_values[index], y_value(float(value))))
    if active:
        segments.append(active)
    return segments


def draw_chart(record: IndicatorEvolution) -> None:
    record.output_path.parent.mkdir(parents=True, exist_ok=True)
    image = Image.new(
        "RGBA",
        (CANVAS_WIDTH, CANVAS_HEIGHT),
        COLORS["white"],
    )
    draw = ImageDraw.Draw(image)

    draw.rounded_rectangle(
        (12, 12, CANVAS_WIDTH - 12, CANVAS_HEIGHT - 12),
        radius=34,
        fill=COLORS["white"],
        outline=COLORS["border"],
        width=4,
    )
    draw.rounded_rectangle(
        (12, 12, CANVAS_WIDTH - 12, 28),
        radius=12,
        fill="#9FC8C4",
    )

    title_font = font(54, bold=True)
    subtitle_font = font(31, semibold=True)
    description_font = font(30)
    direction_font = font(30, bold=True)
    pill_label_font = font(23, semibold=True)
    pill_value_font = font(30, bold=True)

    icon_x, icon_y = 78, 82
    draw.line((icon_x, icon_y + 34, icon_x, icon_y), fill=COLORS["orange"], width=5)
    draw.line((icon_x, icon_y + 34, icon_x + 35, icon_y + 34), fill=COLORS["orange"], width=5)
    draw.rectangle((icon_x + 7, icon_y + 19, icon_x + 13, icon_y + 31), fill=COLORS["orange"])
    draw.rectangle((icon_x + 17, icon_y + 10, icon_x + 23, icon_y + 31), fill=COLORS["orange"])
    draw.rectangle((icon_x + 27, icon_y + 2, icon_x + 33, icon_y + 31), fill=COLORS["orange"])

    title = f"Evolução do indicador — {record.indicator_name}"
    title_lines = wrap_by_pixels(draw, title, title_font, 2250)
    title_y = 68
    for line in title_lines[:2]:
        draw.text((138, title_y), line, font=title_font, fill=COLORS["ink"])
        title_y += 66

    pill = (2490, 62, 3090, 174)
    draw.rounded_rectangle(
        pill,
        radius=22,
        fill=COLORS["plot"],
        outline=COLORS["border"],
        width=3,
    )
    draw.text(
        (2522, 79),
        "REGIÃO FUNCIONAL COMPARADA",
        font=pill_label_font,
        fill=COLORS["muted"],
    )
    draw.text(
        (2522, 117),
        f"RF {record.region_number} (do município)",
        font=pill_value_font,
        fill=COLORS["ink"],
    )

    context_y = max(206, title_y + 6)
    context = (
        f"RF{record.region_number} · {record.dimension_name} · "
        f"{record.municipality_name} · "
        f"{selection_rule_text(record.indicator_rank_2025)}"
    )
    draw.text(
        (78, context_y),
        context,
        font=subtitle_font,
        fill=COLORS["muted"],
    )

    description_y = context_y + 58
    description_lines = wrap_by_pixels(
        draw,
        record.description,
        description_font,
        2080,
    )
    for line in description_lines[:2]:
        draw.text(
            (78, description_y),
            line,
            font=description_font,
            fill=COLORS["muted"],
        )
        description_y += 43
    draw.text(
        (2180, context_y + 58),
        direction_text(record.direction),
        font=direction_font,
        fill=COLORS["muted"],
    )
    separator_y = max(390, description_y + 16)
    draw.line(
        (76, separator_y, CANVAS_WIDTH - 76, separator_y),
        fill=COLORS["border"],
        width=3,
    )

    plot_left = 300
    plot_top = max(500, separator_y + 92)
    plot_right = 2640
    plot_bottom = 1470
    plot_width = plot_right - plot_left
    plot_height = plot_bottom - plot_top

    values = [
        float(value)
        for point in record.points
        for value in (
            point.municipality_value,
            point.regional_median,
            point.state_median,
        )
        if value is not None and math.isfinite(float(value))
    ]
    if not values:
        raise ValueError(f"{record.indicator_id}: série sem valores")

    raw_min = min(values)
    raw_max = max(values)
    observed_span = max(raw_max - raw_min, float.fromhex("0x1.0p-52"))
    visual_magnitude = max(
        abs(raw_min),
        abs(raw_max),
        float.fromhex("0x1.0p-52"),
    )
    visual_span = max(observed_span * 1.55, visual_magnitude * 0.55)
    span_padding = visual_span * 0.16
    center = (raw_min + raw_max) / 2
    minimum = center - visual_span / 2 - span_padding
    maximum = center + visual_span / 2 + span_padding
    span = maximum - minimum

    def y_value(value: float) -> float:
        ratio = (value - minimum) / span
        return plot_top + (1 - ratio) * plot_height

    point_count = len(record.points)
    x_values = [
        (
            plot_left + plot_width / 2
            if point_count == 1
            else plot_left + index * plot_width / (point_count - 1)
        )
        for index in range(point_count)
    ]

    draw.rounded_rectangle(
        (plot_left, plot_top, plot_right, plot_bottom),
        radius=22,
        fill=COLORS["plot"],
        outline=COLORS["plot_border"],
        width=3,
    )
    for ratio in (0, 0.25, 0.5, 0.75, 1):
        grid_y = plot_top + ratio * plot_height
        draw_dashed_line(
            draw,
            ((plot_left, grid_y), (plot_right, grid_y)),
            fill=COLORS["grid"],
            width=3,
            dash=(12, 15),
        )

    tick_font = font(28, semibold=True)
    for ratio in (0, 0.5, 1):
        tick_value = maximum - ratio * span
        tick_text = format_axis_tick(tick_value, record)
        tick_y = plot_top + ratio * plot_height
        tick_box = draw.textbbox((0, 0), tick_text, font=tick_font)
        draw.text(
            (plot_left - 30 - (tick_box[2] - tick_box[0]), tick_y - 17),
            tick_text,
            font=tick_font,
            fill=COLORS["muted_light"],
        )

    y_axis_text = axis_label(record)
    y_axis_font = font(31, semibold=True)
    axis_layer = Image.new("RGBA", (plot_height, 80), (255, 255, 255, 0))
    axis_draw = ImageDraw.Draw(axis_layer)
    axis_box = axis_draw.textbbox((0, 0), y_axis_text, font=y_axis_font)
    axis_draw.text(
        (
            (plot_height - (axis_box[2] - axis_box[0])) / 2,
            18,
        ),
        y_axis_text,
        font=y_axis_font,
        fill=COLORS["muted"],
    )
    rotated_axis = axis_layer.rotate(90, expand=True)
    image.alpha_composite(
        rotated_axis,
        (
            35,
            int(plot_top + (plot_height - rotated_axis.height) / 2),
        ),
    )
    draw = ImageDraw.Draw(image)

    primary_segments = series_segments(
        record.points,
        lambda point: point.municipality_value,
        x_values,
        y_value,
    )
    regional_segments = series_segments(
        record.points,
        lambda point: point.regional_median,
        x_values,
        y_value,
    )
    state_segments = series_segments(
        record.points,
        lambda point: point.state_median,
        x_values,
        y_value,
    )

    area_layer = Image.new("RGBA", image.size, (255, 255, 255, 0))
    area_draw = ImageDraw.Draw(area_layer)
    for segment in primary_segments:
        if len(segment) >= 2:
            polygon = [
                (segment[0][0], plot_bottom),
                *segment,
                (segment[-1][0], plot_bottom),
            ]
            area_draw.polygon(polygon, fill=COLORS["teal_area"])
    image.alpha_composite(area_layer)
    draw = ImageDraw.Draw(image)

    for segment in state_segments:
        draw_dashed_line(
            draw,
            segment,
            fill=COLORS["state"],
            width=8,
            dash=(8, 18),
        )
    for segment in regional_segments:
        draw_dashed_line(
            draw,
            segment,
            fill=COLORS["white"],
            width=15,
            dash=(32, 20),
        )
        draw_dashed_line(
            draw,
            segment,
            fill=COLORS["orange"],
            width=8,
            dash=(32, 20),
        )
    for segment in primary_segments:
        if len(segment) >= 2:
            draw.line(segment, fill=COLORS["teal"], width=13, joint="curve")

    for index, point in enumerate(record.points):
        x_coord = x_values[index]
        if point.state_median is not None:
            y_coord = y_value(float(point.state_median))
            draw.rounded_rectangle(
                (x_coord - 10, y_coord - 10, x_coord + 10, y_coord + 10),
                radius=2,
                fill=COLORS["state"],
                outline=COLORS["white"],
                width=4,
            )
        if point.regional_median is not None:
            y_coord = y_value(float(point.regional_median))
            radius = 13
            draw.polygon(
                (
                    (x_coord, y_coord - radius),
                    (x_coord + radius, y_coord),
                    (x_coord, y_coord + radius),
                    (x_coord - radius, y_coord),
                ),
                fill=COLORS["orange"],
                outline=COLORS["white"],
            )
        if point.municipality_value is not None:
            y_coord = y_value(float(point.municipality_value))
            draw.ellipse(
                (x_coord - 17, y_coord - 17, x_coord + 17, y_coord + 17),
                fill=COLORS["teal"],
                outline=COLORS["white"],
                width=6,
            )

    year_font = font(31, bold=True)
    for index, point in enumerate(record.points):
        label = str(point.data_year)
        box = draw.textbbox((0, 0), label, font=year_font)
        draw.text(
            (
                x_values[index] - (box[2] - box[0]) / 2,
                plot_bottom + 50,
            ),
            label,
            font=year_font,
            fill=COLORS["muted"],
        )

    end_items: list[dict[str, Any]] = []
    for getter, label, tone in (
        (lambda point: point.municipality_value, "Município", "teal"),
        (lambda point: point.regional_median, "Região Funcional", "orange"),
        (lambda point: point.state_median, "RS", "state"),
    ):
        for index in range(len(record.points) - 1, -1, -1):
            value = getter(record.points[index])
            if value is not None:
                target_y = y_value(float(value))
                end_items.append(
                    {
                        "index": index,
                        "label": label,
                        "tone": tone,
                        "target_y": target_y,
                        "label_y": target_y,
                    }
                )
                break
    end_items.sort(key=lambda item: item["target_y"])
    end_gap = 58
    for index in range(1, len(end_items)):
        end_items[index]["label_y"] = max(
            end_items[index]["target_y"],
            end_items[index - 1]["label_y"] + end_gap,
        )
    overflow = max(0, end_items[-1]["label_y"] - (plot_bottom - 25))
    for item in end_items:
        item["label_y"] -= overflow
    underflow = max(0, (plot_top + 25) - end_items[0]["label_y"])
    for item in end_items:
        item["label_y"] += underflow

    end_font = font(30, bold=True)
    for item in end_items:
        tone_color = COLORS[item["tone"]]
        x_start = x_values[item["index"]] + 22
        x_end = plot_right + 52
        draw.line(
            (
                x_start,
                item["target_y"],
                x_end,
                item["label_y"],
            ),
            fill=tone_color,
            width=4,
        )
        draw.text(
            (x_end + 16, item["label_y"] - 20),
            item["label"],
            font=end_font,
            fill=tone_color,
        )

    legend_y = 1740
    legend_font = font(31, semibold=True)
    legend_items = (
        (record.municipality_name, "primary"),
        (f"Mediana da Região Funcional {record.region_number}", "regional"),
        ("Mediana do RS", "state"),
    )
    widths = []
    for label, _ in legend_items:
        widths.append(105 + text_width(draw, label, legend_font))
    total_width = sum(widths) + 130 * (len(widths) - 1)
    legend_x = (CANVAS_WIDTH - total_width) / 2
    for (label, tone), item_width in zip(legend_items, widths):
        line_y = legend_y + 22
        if tone == "primary":
            draw.line(
                (legend_x, line_y, legend_x + 70, line_y),
                fill=COLORS["teal"],
                width=8,
            )
            draw.ellipse(
                (legend_x + 28, line_y - 9, legend_x + 46, line_y + 9),
                fill=COLORS["teal"],
                outline=COLORS["white"],
                width=3,
            )
        elif tone == "regional":
            draw_dashed_line(
                draw,
                ((legend_x, line_y), (legend_x + 70, line_y)),
                fill=COLORS["orange"],
                width=7,
                dash=(25, 14),
            )
            draw.polygon(
                (
                    (legend_x + 37, line_y - 10),
                    (legend_x + 47, line_y),
                    (legend_x + 37, line_y + 10),
                    (legend_x + 27, line_y),
                ),
                fill=COLORS["orange"],
            )
        else:
            draw_dashed_line(
                draw,
                ((legend_x, line_y), (legend_x + 70, line_y)),
                fill=COLORS["state"],
                width=7,
                dash=(7, 15),
            )
            draw.rectangle(
                (legend_x + 29, line_y - 8, legend_x + 45, line_y + 8),
                fill=COLORS["state"],
            )
        draw.text(
            (legend_x + 88, legend_y),
            label,
            font=legend_font,
            fill=COLORS["muted"],
        )
        legend_x += item_width + 130

    image.convert("RGB").save(
        record.output_path,
        format="PNG",
        dpi=(DPI, DPI),
        optimize=True,
    )


EXCEL_COLORS = {
    key: value.removeprefix("#")
    for key, value in COLORS.items()
    if isinstance(value, str) and value.startswith("#")
}
THIN = Side(style="thin", color=EXCEL_COLORS["border"])
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(cell, *, left: bool = False) -> None:
    cell.fill = PatternFill("solid", fgColor=EXCEL_COLORS["header"])
    cell.font = Font(
        name="Aptos Display",
        size=11,
        bold=True,
        color=EXCEL_COLORS["ink"],
    )
    cell.alignment = Alignment(
        horizontal="left" if left else "center",
        vertical="center",
        wrap_text=True,
    )
    cell.border = CELL_BORDER


def style_body(
    cell,
    alternate: bool,
    *,
    left: bool = False,
    bold: bool = False,
    muted: bool = False,
) -> None:
    cell.fill = PatternFill(
        "solid",
        fgColor=EXCEL_COLORS["row_alt"] if alternate else EXCEL_COLORS["white"],
    )
    cell.font = Font(
        name="Aptos",
        size=10.5,
        bold=bold,
        color=EXCEL_COLORS["muted"] if muted else EXCEL_COLORS["ink"],
    )
    cell.alignment = Alignment(
        horizontal="left" if left else "center",
        vertical="center",
        wrap_text=True,
    )
    cell.border = CELL_BORDER


def apply_value_number_format(cell, record: IndicatorEvolution) -> None:
    decimals = "0" * record.decimal_places
    base = "0" if not decimals else f"0.{decimals}"
    if record.value_format == "currency":
        cell.number_format = f'R$ #,##{base}'
    elif record.value_format == "percent" or record.unit == "%":
        cell.number_format = f'{base}"%"'
    else:
        cell.number_format = base


def relative_png(record: IndicatorEvolution) -> str:
    return record.output_path.relative_to(EXPORT_ROOT).as_posix()


def add_summary_sheet(workbook: Workbook, records: list[IndicatorEvolution]) -> None:
    worksheet = workbook.create_sheet("Resumo")
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.zoomScale = 85
    headers = [
        "RF",
        "DIMENSÃO",
        "MUNICÍPIO LÍDER",
        "COREDE",
        "POSIÇÃO NA DIMENSÃO 2025",
        "INDICADOR SELECIONADO",
        "MELHOR POSIÇÃO DO INDICADOR 2025",
        "NOTA DO INDICADOR 2025",
        "DIREÇÃO",
        "UNIDADE",
        "ARQUIVO PNG",
    ]
    for column, header in enumerate(headers, start=1):
        style_header(
            worksheet.cell(row=1, column=column, value=header),
            left=column in (2, 3, 4, 6, 9, 10, 11),
        )
    worksheet.row_dimensions[1].height = 42

    for row_index, record in enumerate(records, start=2):
        direction = {
            "higher_is_better": "Mais alto é melhor",
            "lower_is_better": "Mais baixo é melhor",
        }.get(record.direction, "Neutra")
        values = [
            f"RF{record.region_number}",
            record.dimension_name,
            record.municipality_name,
            record.corede_name,
            record.dimension_rank,
            record.indicator_name,
            record.indicator_rank_2025,
            record.indicator_score_2025,
            direction,
            record.unit,
            relative_png(record),
        ]
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_index, column=column, value=value)
            style_body(
                cell,
                row_index % 2 == 1,
                left=column in (2, 3, 4, 6, 9, 10, 11),
                bold=column in (1, 2, 3, 6, 7),
                muted=column in (4, 9, 10, 11),
            )
            if column in (5, 7):
                cell.number_format = '0"º"'
            elif column == 8:
                cell.number_format = "0.00"
        link_cell = worksheet.cell(row=row_index, column=11)
        link_cell.hyperlink = relative_png(record)
        link_cell.style = "Hyperlink"
        link_cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )
        worksheet.row_dimensions[row_index].height = 35

    worksheet.auto_filter.ref = f"A1:K{len(records) + 1}"
    widths = [9, 20, 29, 28, 19, 48, 20, 18, 22, 18, 68]
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.oddHeader.center.text = (
        "Melhores indicadores dos municípios líderes por dimensão — 2025"
    )
    worksheet.oddHeader.center.font = "Aptos,Bold"
    worksheet.oddFooter.right.text = "Página &P de &N"


def add_region_sheet(
    workbook: Workbook,
    region_number: int,
    records: list[IndicatorEvolution],
) -> None:
    region_records = [
        record for record in records if record.region_number == region_number
    ]
    worksheet = workbook.create_sheet(f"RF{region_number}")
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.zoomScale = 78
    headers = [
        "DIMENSÃO",
        "MUNICÍPIO LÍDER",
        "COREDE",
        "POS. DIMENSÃO 2025",
        "INDICADOR",
        "POS. INDICADOR 2025",
        "ANO NO GRÁFICO",
        "ANO DE REFERÊNCIA",
        "VALOR DO MUNICÍPIO",
        "MEDIANA DA RF",
        "MEDIANA DO RS",
        "POS. ANUAL DO INDICADOR",
        "IMPUTADO",
        "DIREÇÃO",
        "UNIDADE",
        "ARQUIVO PNG",
    ]
    for column, header in enumerate(headers, start=1):
        style_header(
            worksheet.cell(row=1, column=column, value=header),
            left=column in (1, 2, 3, 5, 14, 15, 16),
        )
    worksheet.row_dimensions[1].height = 45

    row_index = 2
    for record in region_records:
        direction = {
            "higher_is_better": "Mais alto é melhor",
            "lower_is_better": "Mais baixo é melhor",
        }.get(record.direction, "Neutra")
        for point in record.points:
            values = [
                record.dimension_name,
                record.municipality_name,
                record.corede_name,
                record.dimension_rank,
                record.indicator_name,
                record.indicator_rank_2025,
                point.data_year,
                point.reference_year,
                point.municipality_value,
                point.regional_median,
                point.state_median,
                point.indicator_rank,
                "Sim" if point.is_imputed else "Não",
                direction,
                record.unit,
                relative_png(record),
            ]
            for column, value in enumerate(values, start=1):
                cell = worksheet.cell(
                    row=row_index,
                    column=column,
                    value=value,
                )
                style_body(
                    cell,
                    row_index % 2 == 1,
                    left=column in (1, 2, 3, 5, 14, 15, 16),
                    bold=column in (1, 2, 5, 6),
                    muted=column in (3, 14, 15, 16),
                )
                if column in (4, 6, 12):
                    cell.number_format = '0"º"'
                elif column in (9, 10, 11):
                    apply_value_number_format(cell, record)
            link_cell = worksheet.cell(row=row_index, column=16)
            link_cell.hyperlink = relative_png(record)
            link_cell.style = "Hyperlink"
            link_cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )
            worksheet.row_dimensions[row_index].height = 31
            row_index += 1

    worksheet.auto_filter.ref = f"A1:P{row_index - 1}"
    widths = [
        20, 28, 27, 18, 48, 18, 16, 17,
        20, 18, 18, 19, 13, 22, 18, 66,
    ]
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.oddHeader.center.text = (
        f"RF{region_number} — evolução dos melhores indicadores — 2021 a 2025"
    )
    worksheet.oddHeader.center.font = "Aptos,Bold"
    worksheet.oddFooter.right.text = "Página &P de &N"


def add_methodology_sheet(
    workbook: Workbook,
    records: list[IndicatorEvolution],
) -> None:
    worksheet = workbook.create_sheet("Metodologia")
    worksheet.sheet_view.showGridLines = False
    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 120
    rows = [
        (
            "Escopo",
            "Regiões Funcionais 1 a 9; seis dimensões; municípios líderes no ranking dimensional de 2025.",
        ),
        (
            "Regra de seleção",
            "Para cada município líder e dimensão, foi identificada a menor posição alcançada por seus indicadores em 2025. Todos os indicadores empatados nessa melhor posição foram incluídos.",
        ),
        (
            "Quantidade",
            f"{len(records)} gráficos, derivados de 54 combinações de Região Funcional e dimensão.",
        ),
        (
            "Série do município",
            "Valor original do indicador do município. O formato, a unidade, o multiplicador e as casas decimais seguem o catálogo da plataforma.",
        ),
        (
            "Mediana da RF",
            "Mediana do valor original na Região Funcional do município, conforme o catálogo da plataforma; quando necessário, usa-se o valor regional presente no arquivo municipal.",
        ),
        (
            "Mediana do RS",
            "Mediana estadual do valor original, conforme o catálogo da plataforma.",
        ),
        (
            "Anos",
            "Referências 2021 a 2025. Quando o catálogo mapeia mais de um ano de referência para o mesmo ano do dado, o gráfico replica o comportamento da plataforma e mostra um único ponto por ano do dado, usando a referência mais recente.",
        ),
        (
            "Leitura",
            "A direção interpretativa indica se valores mais altos ou mais baixos representam melhor desempenho. As posições menores representam melhor colocação no ranking.",
        ),
        (
            "Cores",
            "Município: teal #08716D; mediana da RF: laranja #B86F12; mediana do RS: azul-acinzentado #536B85.",
        ),
    ]
    for row_index, (label, description) in enumerate(rows, start=1):
        label_cell = worksheet.cell(row=row_index, column=1, value=label)
        description_cell = worksheet.cell(
            row=row_index,
            column=2,
            value=description,
        )
        style_header(label_cell, left=True)
        style_body(
            description_cell,
            row_index % 2 == 0,
            left=True,
        )
        worksheet.row_dimensions[row_index].height = 52
    worksheet["A1"].fill = PatternFill(
        "solid",
        fgColor=EXCEL_COLORS["teal"],
    )
    worksheet["A1"].font = Font(
        name="Aptos Display",
        size=11,
        bold=True,
        color=EXCEL_COLORS["white"],
    )


def generate_workbook(records: list[IndicatorEvolution]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    add_summary_sheet(workbook, records)
    for region_number in range(1, 10):
        add_region_sheet(workbook, region_number, records)
    add_methodology_sheet(workbook, records)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(WORKBOOK_PATH)


def create_package(records: list[IndicatorEvolution]) -> None:
    with zipfile.ZipFile(
        ZIP_PATH,
        mode="w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=7,
    ) as archive:
        archive.write(WORKBOOK_PATH, WORKBOOK_PATH.name)
        for record in records:
            archive.write(
                record.output_path,
                relative_png(record),
            )


def validate_outputs(records: list[IndicatorEvolution]) -> None:
    if len(records) != 140:
        raise ValueError(
            f"Esperados 140 gráficos pela seleção validada; encontrados {len(records)}"
        )
    paths = [record.output_path for record in records]
    if len(paths) != len(set(paths)):
        raise ValueError("Há nomes de PNG duplicados")
    for path in paths:
        if not path.exists():
            raise FileNotFoundError(path)
        with Image.open(path) as image:
            if image.size != (CANVAS_WIDTH, CANVAS_HEIGHT):
                raise ValueError(f"{path.name}: tamanho {image.size}")
            if image.info.get("dpi"):
                dpi_x, dpi_y = image.info["dpi"]
                if abs(dpi_x - DPI) > 1 or abs(dpi_y - DPI) > 1:
                    raise ValueError(f"{path.name}: DPI {image.info['dpi']}")

    workbook = load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
    expected_sheets = ["Resumo", *[f"RF{i}" for i in range(1, 10)], "Metodologia"]
    if workbook.sheetnames != expected_sheets:
        raise ValueError(f"Abas inesperadas: {workbook.sheetnames}")
    if workbook["Resumo"].max_row != len(records) + 1:
        raise ValueError("Resumo com quantidade incorreta de linhas")
    for region_number in range(1, 10):
        expected_rows = (
            sum(
                len(record.points)
                for record in records
                if record.region_number == region_number
            )
            + 1
        )
        if workbook[f"RF{region_number}"].max_row != expected_rows:
            raise ValueError(
                f"RF{region_number}: linhas inesperadas no Excel"
            )
    workbook.close()

    if not ZIP_PATH.exists():
        raise FileNotFoundError(ZIP_PATH)
    with zipfile.ZipFile(ZIP_PATH) as archive:
        expected = {WORKBOOK_PATH.name, *[relative_png(record) for record in records]}
        if set(archive.namelist()) != expected:
            raise ValueError("Conteúdo do ZIP não corresponde aos artefatos")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Gera os gráficos de evolução dos melhores indicadores dos "
            "municípios líderes por dimensão e o Excel correspondente."
        )
    )
    parser.add_argument(
        "--regions",
        nargs="*",
        type=int,
        default=list(range(1, 10)),
        help="Regiões Funcionais a gerar (padrão: 1 a 9).",
    )
    parser.add_argument(
        "--clean",
        action="store_true",
        help="Remove a pasta de saída antes de gerar.",
    )
    parser.add_argument(
        "--skip-validation",
        action="store_true",
        help="Não executa a validação final.",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=4,
        help="Quantidade de PNGs renderizados em paralelo (padrão: 4).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    regions = sorted(set(args.regions))
    if any(region not in range(1, 10) for region in regions):
        raise ValueError("As Regiões Funcionais válidas são 1 a 9")
    if args.clean and EXPORT_ROOT.exists():
        shutil.rmtree(EXPORT_ROOT)
    records = build_records(region_numbers=regions)
    print(f"Registros selecionados: {len(records)}")
    if args.workers < 1:
        raise ValueError("--workers deve ser pelo menos 1")
    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        for index, _ in enumerate(
            executor.map(draw_chart, records),
            start=1,
        ):
            if index % 10 == 0 or index == len(records):
                print(f"PNG {index}/{len(records)}")

    if regions == list(range(1, 10)):
        generate_workbook(records)
        create_package(records)
        if not args.skip_validation:
            validate_outputs(records)
        print(f"Excel: {WORKBOOK_PATH}")
        print(f"ZIP: {ZIP_PATH}")
    else:
        print("Geração parcial: Excel e ZIP não foram recriados.")


if __name__ == "__main__":
    main()
