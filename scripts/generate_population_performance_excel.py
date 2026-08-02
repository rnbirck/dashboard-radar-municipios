from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_ROOT = PROJECT_ROOT / "public" / "data" / "v2025" / "rankings" / "2025"
EXPORT_ROOT = PROJECT_ROOT / "docs" / "exports" / "graficos 3"
OUTPUT_PATH = (
    EXPORT_ROOT
    / "desempenho-populacional-regioes-funcionais-2025.xlsx"
)

COLORS = {
    "ink": "142A41",
    "muted": "52657A",
    "header": "E9F2F6",
    "border": "D8E2E9",
    "row_alt": "F8FAFC",
    "white": "FFFFFF",
    "rank_fill": "E9EEF3",
    "rank_text": "344A5F",
    "above_fill": "72C4A5",
    "above_stroke": "087654",
    "expected_fill": "E6B84B",
    "expected_stroke": "8A5A12",
    "below_fill": "E9898E",
    "below_stroke": "B4232D",
}

CLASSIFICATIONS = {
    "above": ("Acima", 1, "above_fill", "above_stroke"),
    "expected": ("No intervalo", 0, "expected_fill", "expected_stroke"),
    "below": ("Abaixo", -1, "below_fill", "below_stroke"),
}

THIN = Side(style="thin", color=COLORS["border"])
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_region(region_number: int) -> dict:
    path = DATA_ROOT / f"rf{region_number}.json"
    return json.loads(path.read_text(encoding="utf-8"))["data"]


def style_header(cell, *, left: bool = False) -> None:
    cell.fill = PatternFill("solid", fgColor=COLORS["header"])
    cell.font = Font(
        name="Aptos Display",
        size=11,
        bold=True,
        color=COLORS["ink"],
    )
    cell.alignment = Alignment(
        horizontal="left" if left else "center",
        vertical="center",
        wrap_text=True,
    )
    cell.border = CELL_BORDER


def style_identity(cell, alternate: bool, *, muted: bool = False) -> None:
    cell.fill = PatternFill(
        "solid",
        fgColor=COLORS["row_alt"] if alternate else COLORS["white"],
    )
    cell.font = Font(
        name="Aptos",
        size=11,
        bold=True,
        color=COLORS["muted"] if muted else COLORS["ink"],
    )
    cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )
    cell.border = CELL_BORDER


def style_rank(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=COLORS["rank_fill"])
    cell.font = Font(
        name="Aptos",
        size=11,
        bold=True,
        color=COLORS["rank_text"],
    )
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = '0"º"'
    cell.border = CELL_BORDER


def style_classification(cell, code: str) -> None:
    _, _, fill_key, stroke_key = CLASSIFICATIONS[code]
    stroke = COLORS[stroke_key]
    cell.fill = PatternFill("solid", fgColor=COLORS[fill_key])
    cell.font = Font(
        name="Aptos",
        size=11,
        bold=True,
        color=stroke,
    )
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin", color=stroke),
        right=Side(style="thin", color=stroke),
        top=Side(style="thin", color=stroke),
        bottom=Side(style="thin", color=stroke),
    )


def add_region_sheet(workbook: Workbook, region_number: int) -> None:
    region = load_region(region_number)
    municipalities = sorted(
        region["municipalities"],
        key=lambda item: item["municipalityName"].casefold(),
    )

    worksheet = workbook.create_sheet(title=f"RF{region_number}")
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 90
    worksheet.freeze_panes = "A2"

    headers = [
        "MUNICÍPIO",
        "COREDE",
        "POSIÇÃO GERAL 2025",
        "CLASSIFICAÇÃO",
        "VALOR NO GRÁFICO",
    ]
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column, value=header)
        style_header(cell, left=column in (1, 2))
    worksheet.row_dimensions[1].height = 34

    for row_index, municipality in enumerate(municipalities, start=2):
        alternate = row_index % 2 == 1
        worksheet.row_dimensions[row_index].height = 28

        municipality_cell = worksheet.cell(
            row=row_index,
            column=1,
            value=municipality["municipalityName"],
        )
        corede_cell = worksheet.cell(
            row=row_index,
            column=2,
            value=municipality["coredeName"],
        )
        rank_cell = worksheet.cell(
            row=row_index,
            column=3,
            value=int(municipality["overallRank"]),
        )

        code = municipality["populationPerformance"]["code"]
        label, chart_value, _, _ = CLASSIFICATIONS[code]
        classification_cell = worksheet.cell(
            row=row_index,
            column=4,
            value=label,
        )
        chart_value_cell = worksheet.cell(
            row=row_index,
            column=5,
            value=chart_value,
        )

        style_identity(municipality_cell, alternate)
        style_identity(corede_cell, alternate, muted=True)
        style_rank(rank_cell)
        style_classification(classification_cell, code)
        style_classification(chart_value_cell, code)
        chart_value_cell.number_format = '+0;-0;0'

    last_row = len(municipalities) + 1
    worksheet.auto_filter.ref = f"A1:E{last_row}"
    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 31
    worksheet.column_dimensions["C"].width = 20
    worksheet.column_dimensions["D"].width = 20
    worksheet.column_dimensions["E"].width = 20

    chart_path = (
        EXPORT_ROOT
        / f"desempenho-populacional-rf{region_number}-2025.png"
    )
    if not chart_path.exists():
        raise FileNotFoundError(f"Gráfico ausente: {chart_path}")

    chart_image = Image(chart_path)
    original_width = chart_image.width
    original_height = chart_image.height
    chart_image.width = 1250
    chart_image.height = round(
        original_height * chart_image.width / original_width
    )
    worksheet.add_image(chart_image, "G2")

    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.35
    worksheet.page_margins.bottom = 0.35
    worksheet.oddHeader.center.text = (
        f"Região Funcional {region_number} — desempenho populacional em 2025"
    )
    worksheet.oddHeader.center.font = "Aptos,Bold"
    worksheet.oddHeader.center.size = 11
    worksheet.oddFooter.right.text = "Página &P de &N"


def validate_workbook(path: Path) -> None:
    workbook = load_workbook(path, data_only=False)
    expected_sheets = [f"RF{region_number}" for region_number in range(1, 10)]
    if workbook.sheetnames != expected_sheets:
        raise ValueError(
            f"Abas inesperadas: {workbook.sheetnames}; "
            f"esperado: {expected_sheets}"
        )

    for region_number, sheet_name in enumerate(expected_sheets, start=1):
        worksheet = workbook[sheet_name]
        region = load_region(region_number)
        expected_municipalities = sorted(
            region["municipalities"],
            key=lambda item: item["municipalityName"].casefold(),
        )

        if worksheet.max_row != len(expected_municipalities) + 1:
            raise ValueError(
                f"{sheet_name}: quantidade de linhas inesperada "
                f"({worksheet.max_row})"
            )
        if worksheet.max_column != 5:
            raise ValueError(
                f"{sheet_name}: quantidade de colunas inesperada "
                f"({worksheet.max_column})"
            )
        if worksheet["A1"].value != "MUNICÍPIO":
            raise ValueError(f"{sheet_name}: cabeçalho de município inválido")
        if worksheet["E1"].value != "VALOR NO GRÁFICO":
            raise ValueError(f"{sheet_name}: último cabeçalho inválido")
        if len(worksheet._images) != 1:
            raise ValueError(
                f"{sheet_name}: esperado 1 gráfico incorporado, "
                f"encontrados {len(worksheet._images)}"
            )

        for row_index, municipality in enumerate(
            expected_municipalities,
            start=2,
        ):
            actual_name = worksheet.cell(row=row_index, column=1).value
            if actual_name != municipality["municipalityName"]:
                raise ValueError(
                    f"{sheet_name}: ordem alfabética divergente na linha "
                    f"{row_index}: {actual_name}"
                )

            code = municipality["populationPerformance"]["code"]
            expected_label, expected_value, _, _ = CLASSIFICATIONS[code]
            actual_label = worksheet.cell(row=row_index, column=4).value
            actual_value = worksheet.cell(row=row_index, column=5).value
            if (actual_label, actual_value) != (
                expected_label,
                expected_value,
            ):
                raise ValueError(
                    f"{sheet_name} {actual_name}: classificação divergente"
                )

    workbook.close()


def main() -> None:
    EXPORT_ROOT.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = (
        "Desempenho dos municípios em relação ao porte populacional — 2025"
    )
    workbook.properties.subject = (
        "Classificação dos municípios das Regiões Funcionais 1 a 9"
    )
    workbook.properties.creator = "Codex"
    workbook.properties.description = (
        "Município, Corede, posição geral de 2025, classificação e valor "
        "utilizado nos gráficos, em ordem alfabética."
    )

    for region_number in range(1, 10):
        add_region_sheet(workbook, region_number)

    workbook.save(OUTPUT_PATH)
    validate_workbook(OUTPUT_PATH)
    print(f"Excel gerado e validado: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
